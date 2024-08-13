"""
Description:
This script automates network management and data collection tasks. 
It connects to network devices such as routers and switches, retrieves various configurations and statuses 
(e.g., VLAN configurations, ARP tables, MAC addresses, interface statuses, etc.), and exports the collected data to an Excel file. 
The script uses the Netmiko library for SSH connections, TextFSM for parsing command outputs, and OpenPyXL for handling Excel files. 
Additional functionalities include multithreading for concurrent device access, data merging, filtering, sorting, and applying formatting to the Excel output.

Authorship:
Developed by: Sebastian Skubisz
Email: sskubisz9@gmail.com
Date Created: June 10th, 2024
Last Modified: August 8th, 2024

Changelog:
- Added functionality to collect VLAN advanced configurations.
- Improved data parsing and merging logic.
- Implemented data filtering, sorting, and deduplication.
- Enhanced Excel export with formatting and additional columns.
- Integrated multithreading for faster data collection.

"""

import time
import datetime
import logging
import os
from netmiko import ConnectHandler, NetMikoTimeoutException, NetMikoAuthenticationException
import pandas as pd
import textfsm
import concurrent.futures
import pprint
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import keyring
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess

# Set the time format for logging and file names
TNOW = datetime.datetime.now()
TFORMAT = '{:%m-%d-%Y_%Hh-%Mm-%Ss}'.format(TNOW)

# Determine the base path based on the script's location
BASE_PATH = os.path.dirname(os.path.abspath(__file__))

# Get credentials from environment variables
username = keyring.get_password("network", "username")
password = keyring.get_password("network", "password")
enable_pass = keyring.get_password("network", "enable_pass")

# Paths to the templates and device lists
TEMPLATE_PATH_ROUTE = os.path.join(BASE_PATH, 'ntc-templates-master/ntc-templates-master/ntc_templates/templates/extreme_ers_show_ip_route_vrfid.textfsm')
TEMPLATE_PATH_VLAN_ADVANCE = os.path.join(BASE_PATH, 'ntc-templates-master/ntc-templates-master/ntc_templates/templates/extreme_ers_show_vlan_advance.textfsm')
ROUTER_LIST_PATH = os.path.join(BASE_PATH, 'Router.txt')
SWITCH_LIST_PATH = os.path.join(BASE_PATH, 'Switch.txt')
TEMPLATE_PATH_VRF = os.path.join(BASE_PATH, 'ntc-templates-master/ntc-templates-master/ntc_templates/templates/extreme_ers_show_ip_vrf_id_only.textfsm')
TEMPLATE_PATH_ARP = os.path.join(BASE_PATH, 'ntc-templates-master/ntc-templates-master/ntc_templates/templates/extreme_ers_show_ip_arp_vrfid.textfsm')
TEMPLATE_PATH_INTERFACE = os.path.join(BASE_PATH, 'ntc-templates-master/ntc-templates-master/ntc_templates/templates/extreme_ers_show_interface_name.textfsm')
TEMPLATE_PATH_MAC = os.path.join(BASE_PATH, 'ntc-templates-master/ntc-templates-master/ntc_templates/templates/extreme_ers_show_mac-address-table.textfsm')
TEMPLATE_PATH_PORT_STATUS = os.path.join(BASE_PATH, 'ntc-templates-master/ntc-templates-master/ntc_templates/templates/extreme_ers_show_interfaces.textfsm')
TEMPLATE_PATH_PING = os.path.join(BASE_PATH, 'ntc-templates-master/ntc-templates-master/ntc_templates/templates/extreme_ers_ping.textfsm')
TEMPLATE_PATH_VLAN = os.path.join(BASE_PATH, 'ntc-templates-master/ntc-templates-master/ntc_templates/templates/extreme_ers_show_running_config_vlan.textfsm')
VRF_ID_OUTPUT_PATH = os.path.join(BASE_PATH, 'Vrf_List.txt')
EXCEL_OUTPUT_PATH = os.path.join(BASE_PATH, f'Network_Scraper_Output_{TFORMAT}.xlsx')

# Read router and switch lists
with open(ROUTER_LIST_PATH, 'r') as f:
    router_list = [line.strip() for line in f if line.strip()]

with open(SWITCH_LIST_PATH, 'r') as f:
    switch_list = [line.strip() for line in f if line.strip()]

vrf_ids = set()
all_data = []
mac_table = []
port_list = []
arp_data = []
port_status_list = []
ping_results = []
vlan_configurations = []  # New list for VLAN configurations
vrf_entries = []  # List to store VRF_NAME and VRF_ID
vlan_advance_data = []  # List to store VLAN advance data

def parse_textfsm_output(output, template_path):
    with open(template_path) as template_file:
        fsm = textfsm.TextFSM(template_file)
        parsed_output = fsm.ParseText(output)
    return [dict(zip(fsm.header, entry)) for entry in parsed_output]

def collect_vrf_id_info(net_connect, rtr):
    output = net_connect.send_command('show ip vrf')
    vrf_data = parse_textfsm_output(output, TEMPLATE_PATH_VRF)
    for data in vrf_data:
        vrf_ids.add(data['VRF_ID'])
        data['Device'] = rtr
        vrf_entries.append({'VRF_NAME': data['VRF_NAME'], 'VRF_ID': data['VRF_ID']})
        all_data.append(data)

    with open(VRF_ID_OUTPUT_PATH, 'w') as vrf_file:
        for vrf_id in sorted(vrf_ids):
            vrf_file.write(f'{vrf_id}\n')

def collect_arp_info(net_connect, rtr):
    global arp_data  # Declare arp_data as global to modify it inside the function
    with open(VRF_ID_OUTPUT_PATH, 'r') as f:
        vrf_list = f.readlines()

    for vrf_id in vrf_list:
        vrf_id = vrf_id.strip()
        arp_output = net_connect.send_command(f'show ip arp vrfid {vrf_id}')
        print(f"\nDebug: Raw ARP Output for {rtr} (VRF {vrf_id}):\n{arp_output}")
        arp_entries = parse_textfsm_output(arp_output, TEMPLATE_PATH_ARP)
        print(f"\nDebug: Parsed ARP Entries for {rtr} (VRF {vrf_id}):")
        pprint.pprint(arp_entries)
        for entry in arp_entries:
            entry['VRF_ID'] = vrf_id  # Add VRF_ID to each ARP entry
        arp_data.extend(arp_entries)

def collect_mac_info(net_connect):
    mac_output = net_connect.send_command('show mac-address-table')
    mac_table.extend(parse_textfsm_output(mac_output, TEMPLATE_PATH_MAC))

def collect_interface_info(net_connect):
    port_output = net_connect.send_command('show interface name')
    port_list.extend(parse_textfsm_output(port_output, TEMPLATE_PATH_INTERFACE))

def collect_port_status_info(net_connect):
    port_status_output = net_connect.send_command('show interfaces')
    port_status_entries = parse_textfsm_output(port_status_output, TEMPLATE_PATH_PORT_STATUS)
    
    # Process entries to separate UNIT and PORT
    for entry in port_status_entries:
        unit_port = entry.get('UNIT_PORT', '')
        if unit_port:
            unit_parts = unit_port.split('/')
            if len(unit_parts) == 2:
                entry['UNIT'] = unit_parts[0]
                entry['PORT'] = unit_parts[1]
            else:
                # Handle cases where UNIT might not be in the expected format
                entry['UNIT'] = ''
                entry['PORT'] = unit_port
    
    port_status_list.extend(port_status_entries)



def collect_vlan_configurations(net_connect):
    global vlan_configurations
    vlan_output = net_connect.send_command('show running-config module vlan')
    vlan_entries = parse_textfsm_output(vlan_output, TEMPLATE_PATH_VLAN)
    for entry in vlan_entries:
        entry['PREFIX'] = None  # Add placeholder for PREFIX column
        vlan_configurations.append(entry)

def collect_vlan_advance(net_connect, rtr):
    global vlan_advance_data
    vlan_output = net_connect.send_command('show vlan advance')
    vlan_entries = parse_textfsm_output(vlan_output, TEMPLATE_PATH_VLAN_ADVANCE)
    vlan_advance_data.extend(vlan_entries)

def filter_and_sort_vlans(vlan_config_list):
    # Remove duplicates by converting to a set of tuples
    unique_vlans = {tuple(vlan.items()) for vlan in vlan_config_list}
    
    # Convert back to list of dictionaries
    filtered_vlans = [dict(vlan) for vlan in unique_vlans]
    
    # Sort by VLAN_ID
    sorted_vlans = sorted(filtered_vlans, key=lambda x: int(x['VLAN_ID']))
    
    return sorted_vlans

def mask_to_prefix(mask):
    return sum(bin(int(octet)).count('1') for octet in mask.split('.'))

def add_prefix_column(vlan_config_list):
    for vlan in vlan_config_list:
        subnet_mask = vlan.get('SUBNET_MASK', None)
        if subnet_mask:
            prefix_length = mask_to_prefix(subnet_mask)

            vlan['PREFIX'] = f'/{prefix_length}'
       

        else:
            vlan['PREFIX'] = None

        # Move PREFIX to be after IP
        if 'IP' in vlan:
            items = list(vlan.items())
            ip_index = [i for i, (k, v) in enumerate(items) if k == 'IP'][0]
            vlan_items = items[:ip_index + 1] + [('PREFIX', vlan['PREFIX'])] + items[ip_index + 1:]
            vlan.clear()
            vlan.update(vlan_items)
    return vlan_config_list

def ping_ips(final_merged_list):
    global ping_results  # Declare ping_results as global to modify it inside the function
    
    for entry in final_merged_list:
        ip_address = entry.get('IP_ADDRESS')
        if ip_address:
            try:
                # Ping the IP address with 4 packets and a timeout of 5 seconds
                response = subprocess.run(['ping', '-n', '4', ip_address], capture_output=True, text=True, timeout=5)
                
                # Check if any packets were lost
                if "Received = 4" in response.stdout:
                    ping_results.append({'IP_ADDRESS': ip_address, 'STATUS': 'Good'})
                else:
                    ping_results.append({'IP_ADDRESS': ip_address, 'STATUS': 'Bad'})
            
            except subprocess.TimeoutExpired:
                print(f"Ping to {ip_address} timed out.")
                ping_results.append({'IP_ADDRESS': ip_address, 'STATUS': 'Bad'})
            
            except Exception as e:
                print(f"An error occurred while pinging {ip_address}: {e}")
                ping_results.append({'IP_ADDRESS': ip_address, 'STATUS': 'Bad'})

            # Debug output to track progress
            print(f"Completed pinging {ip_address}, status recorded.")


def backup_device(rtr, device_type):
    device = {
        "device_type": "avaya_ers",  # Adjust based on your device type
        "host": rtr,
        "username": username,
        "password": password,
        "secret": enable_pass,
        "port": 22,
        "verbose": True,
        "session_log": f'log_{rtr}.txt'
    }

    try:
        net_connect = ConnectHandler(**device)
        net_connect.enable()

        # Send commands to set terminal settings
        net_connect.send_command('terminal length 0')
        net_connect.send_command('terminal more disable')
        net_connect.send_command('disable clipaging')
        net_connect.send_command('en')

        # Collect data based on device type
        if device_type == 'router':
            collect_vrf_id_info(net_connect, rtr)
            collect_arp_info(net_connect, rtr)
            collect_vlan_configurations(net_connect)  # Collect VLAN info for routers
            collect_vlan_advance(net_connect, rtr)  # Collect VLAN advance info
        elif device_type == 'switch':
            collect_mac_info(net_connect)
            collect_interface_info(net_connect)
            collect_port_status_info(net_connect)

        net_connect.disconnect()
        logging.info(f'Backup of {rtr} completed successfully.')
        print(f'Backup of {rtr} completed successfully.')

    except (NetMikoTimeoutException, NetMikoAuthenticationException) as e:
        logging.error(f"Error: Access to {rtr} failed, backup was not taken. Exception: {str(e)}")
        print(f'Error: Access to {rtr} failed, backup was not taken')
    except Exception as e:
        logging.error(f"Error: An unexpected error occurred with {rtr}. Exception: {str(e)}")
        print(f'Error: An unexpected error occurred with {rtr}. Exception: {str(e)}')

# Use multithreading for concurrent execution
with concurrent.futures.ThreadPoolExecutor(max_workers=100) as executor:
    futures = []
    for router in router_list:
        futures.append(executor.submit(backup_device, router, 'router'))
    for switch in switch_list:
        futures.append(executor.submit(backup_device, switch, 'switch'))
    concurrent.futures.wait(futures)

def normalize_mac(mac_address):
    return mac_address.replace(":", "-").lower()

def merge_mac_and_port_tables(mac_table, port_list):
    merged_list = []

    for port_entry in port_list:
        port_merged = False
        for mac_entry in mac_table:
            if port_entry['PORT'] == mac_entry['PORT'] and port_entry.get('UNIT') == mac_entry.get('UNIT'):
                merged_entry = {
                    'UNIT': port_entry.get('UNIT', ''),
                    'PORT': port_entry['PORT'],
                    'NAME': port_entry['NAME'],
                    'VLAN': mac_entry['VID'],
                    'MAC': normalize_mac(mac_entry['MAC_ADDRESS']),
                    'IP_ADDRESS': None,  # Default to None
                    'OPER': None,  # Default to None
                    'SPEED': None,  # Default to None
                    'PING_STATUS': '',  # Default to empty
                    'VRF_ID': None  # Default to None
                }
                merged_list.append(merged_entry)
                port_merged = True
        if not port_merged:
            merged_entry = {
                'UNIT': port_entry.get('UNIT', ''),
                'PORT': port_entry['PORT'],
                'NAME': port_entry['NAME'],
                'VLAN': None,
                'MAC': None,
                'IP_ADDRESS': None,
                'OPER': None,
                'SPEED': None,
                'PING_STATUS': '',  # Default to empty
                'VRF_ID': None  # Default to None
            }
            merged_list.append(merged_entry)

    return merged_list

def merge_with_arp_table(merged_list, arp_data):
    for entry in merged_list:
        if entry['MAC']:
            for arp_entry in arp_data:
                if normalize_mac(entry['MAC']) == normalize_mac(arp_entry['MAC_ADDRESS']):
                    entry['IP_ADDRESS'] = arp_entry['IP_ADDRESS']
                    entry['VRF_ID'] = arp_entry['VRF_ID']
                    break

    return merged_list

def merge_with_port_status(merged_list, port_status_list):
    for entry in merged_list:
        for port_status in port_status_list:
            if entry['UNIT'] == port_status['UNIT'] and entry['PORT'] == port_status['PORT']:
                entry['OPER'] = port_status.get('OPER_STATUS', None)
                entry['SPEED'] = port_status.get('SPEED', None)
                break

    return merged_list

def merge_with_ping_results(merged_list, ping_results):
    # Create a dictionary for fast lookups of ping results by IP address
    ping_dict = {entry['IP_ADDRESS']: entry.get('STATUS', '') for entry in ping_results}
    
    # Update the merged list with ping statuses
    for entry in merged_list:
        ip_address = entry.get('IP_ADDRESS')
        if ip_address:
            entry['PING_STATUS'] = ping_dict.get(ip_address, '')
    
    return merged_list

# Ping IPs after collecting data from devices
# Example usage of ping_ips:
merged_list = merge_mac_and_port_tables(mac_table, port_list)
merged_list_with_ips_and_vrf = merge_with_arp_table(merged_list, arp_data)
final_merged_list = merge_with_port_status(merged_list_with_ips_and_vrf, port_status_list)

# 2. Now, only ping the IPs that are in the final merged list
ping_ips(final_merged_list)

# 3. Merge the ping results back into the final list
final_list_with_pings = merge_with_ping_results(final_merged_list, ping_results)

# Export the final merged list to an Excel file
print(f"\nDebug: Final List with Pings and VRF_IDs:")
pprint.pprint(final_list_with_pings)

df = pd.DataFrame(final_list_with_pings)
df.to_excel(EXCEL_OUTPUT_PATH, index=False)

# Filter and sort VLAN configurations
filtered_sorted_vlans = filter_and_sort_vlans(vlan_configurations)
vlan_configurations_with_prefix = add_prefix_column(filtered_sorted_vlans)

# Prepare VLAN DataFrame
vlan_df = pd.DataFrame(vlan_configurations_with_prefix)

# Filter and sort VRF entries
unique_vrf_entries = {tuple(vrf.items()) for vrf in vrf_entries}
filtered_vrf_entries = [dict(vrf) for vrf in unique_vrf_entries]
sorted_vrf_entries = sorted(filtered_vrf_entries, key=lambda x: int(x['VRF_ID']))

# Prepare VRF DataFrame
vrf_df = pd.DataFrame(sorted_vrf_entries)

# Filter and sort VLAN advance entries
unique_vlan_advance_entries = {tuple(vlan.items()) for vlan in vlan_advance_data}
filtered_vlan_advance_entries = [dict(vlan) for vlan in unique_vlan_advance_entries]
sorted_vlan_advance_entries = sorted(filtered_vlan_advance_entries, key=lambda x: int(x['VLAN_ID']))

# Prepare VLAN advance DataFrame
vlan_advance_df = pd.DataFrame(sorted_vlan_advance_entries)

# Function to append data to Excel starting from a specific column
def append_to_excel(excel_path, df, start_column):
    wb = load_workbook(excel_path)
    ws = wb.active

    # Determine the starting row
    start_row = 1  # Start appending from the second row (to skip headers)

    for row in dataframe_to_rows(df, index=False, header=True):
        for idx, value in enumerate(row):
            ws.cell(row=start_row, column=start_column + idx, value=value)
        start_row += 1

    wb.save(excel_path)

# Append VLAN configurations and VRF entries to the Excel file
append_to_excel(EXCEL_OUTPUT_PATH, vlan_df, 12)
append_to_excel(EXCEL_OUTPUT_PATH, vrf_df, 20)
append_to_excel(EXCEL_OUTPUT_PATH, vlan_advance_df, 23)

# Apply colors and borders to OPER and PING_STATUS columns, and add separators
def apply_colors_and_borders_to_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply colors to OPER and PING_STATUS columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border  # Apply thin border to all cells
            if cell.column_letter == 'G':  # Assuming OPER is in column G
                if cell.value == 'Up':
                    cell.fill = green_fill
                elif cell.value == 'Down':
                    cell.fill = red_fill
            elif cell.column_letter == 'I':  # Assuming PING_STATUS is in column I
                if cell.value == 'Good':
                    cell.fill = green_fill
                elif cell.value == 'Bad':
                    cell.fill = red_fill

    # Make the whole cell black for columns K, S, and V
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for col_letter in ['K', 'S', 'V']:
            cell = row[ws[col_letter][0].column - 1]
            cell.fill = black_fill

    # Adjust column widths for readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Apply additional formatting
    font = Font(name='Calibri', size=11)
    alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment

    # Make header row bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.border = Border(left=Side(style='medium'),
                             right=Side(style='medium'),
                             top=Side(style='medium'),
                             bottom=Side(style='medium'))

    wb.save(excel_path)

# Apply colors and formatting after all data has been appended
apply_colors_and_borders_to_excel(EXCEL_OUTPUT_PATH)
print(f"Data successfully exported to {EXCEL_OUTPUT_PATH}")

# Log the final merged data
print("\nFinal Merged Data:")
pprint.pprint(final_list_with_pings)

"""
Reflections:
Throughout my internship, developing this script has been an invaluable learning experience. 
I gained a deep understanding of network automation, from establishing secure SSH connections using Netmiko to parsing complex command outputs with TextFSM. 
Working with data in Python, I honed my skills in data manipulation and analysis, particularly with pandas and OpenPyXL. 
The project also taught me the importance of efficient coding practices, such as multithreading for performance optimization, 
and meticulous attention to detail in ensuring data accuracy and presentation. 
This journey has not only enhanced my technical skills but also reinforced the significance of problem-solving, perseverance, 
and continuous learning in the field of network management and automation.

-Sebastian Skubisz
"""
